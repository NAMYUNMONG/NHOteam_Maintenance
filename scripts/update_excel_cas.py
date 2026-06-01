#!/usr/bin/env python3
"""
update_excel_cas.py — Write CAS numbers from inventory.json back to Excel

Usage:
    python scripts/update_excel_cas.py data/inventory.json source/Inventory_Master_V7.xlsx

- Matches rows by Item_ID
- Writes CAS_No from JSON into the CAS_No column in Excel
- Never modifies any other column
- Saves Excel in place (overwrites)
- Prints a report of what changed

Requirements:
    pip install openpyxl
"""

import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl")


def main():
    if len(sys.argv) < 3:
        print(__doc__); raise SystemExit(1)

    json_path  = Path(sys.argv[1])
    excel_path = Path(sys.argv[2])

    if not json_path.exists():
        raise SystemExit(f'JSON not found: {json_path}')
    if not excel_path.exists():
        raise SystemExit(f'Excel not found: {excel_path}')

    # Load JSON — build Item_ID → CAS_No lookup
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    cas_map = {
        i['Item_ID']: i.get('CAS_No', '').strip()
        for i in data['items']
        if i.get('CAS_No', '').strip()
    }
    print(f'JSON: {len(cas_map)} items with CAS_No')

    # Load Excel
    wb = openpyxl.load_workbook(excel_path)

    # Find Inventory sheet
    sheet_name = 'Inventory'
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}')
    ws = wb[sheet_name]

    # Find column indices from header row
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    if 'Item_ID' not in headers:
        raise SystemExit(f'Item_ID column not found in header: {list(headers.keys())}')
    if 'CAS_No' not in headers:
        raise SystemExit(f'CAS_No column not found in header: {list(headers.keys())}')

    id_col  = headers['Item_ID']
    cas_col = headers['CAS_No']
    print(f'Excel: Item_ID col={id_col}, CAS_No col={cas_col}')

    # Update CAS_No for matching rows
    updated = 0
    skipped_has_cas = 0
    not_in_json = 0

    for row in ws.iter_rows(min_row=2):
        item_id_cell = row[id_col - 1]
        cas_cell     = row[cas_col - 1]
        item_id = str(item_id_cell.value or '').strip()

        if not item_id:
            continue

        existing_cas = str(cas_cell.value or '').strip()

        if item_id not in cas_map:
            not_in_json += 1
            continue

        new_cas = cas_map[item_id]

        if existing_cas == new_cas:
            continue  # already matches

        if existing_cas and existing_cas != new_cas:
            # Excel already has a different CAS — don't overwrite, just report
            print(f'  CONFLICT {item_id}: Excel={existing_cas} JSON={new_cas} — keeping Excel value')
            skipped_has_cas += 1
            continue

        # Write new CAS
        cas_cell.value = new_cas
        updated += 1

    wb.save(excel_path)

    print(f'\n{"="*55}')
    print(f'  Excel updated: {excel_path.name}')
    print(f'  CAS written  : {updated}')
    print(f'  Conflicts    : {skipped_has_cas} (Excel value kept)')
    print(f'  Not in JSON  : {not_in_json}')
    print(f'{"="*55}\n')


if __name__ == '__main__':
    main()
