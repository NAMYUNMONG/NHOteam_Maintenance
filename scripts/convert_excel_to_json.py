#!/usr/bin/env python3
"""
convert_excel_to_json.py — Convert Inventory_Master_V*.xlsx to data/inventory.json

Usage:
    python scripts/convert_excel_to_json.py <input.xlsx> <output.json>

Example:
    python scripts/convert_excel_to_json.py source/Inventory_Master_V7.xlsx data/inventory.json

Requirements:
    pip install openpyxl

Notes:
    - Reads a single 'Inventory' sheet (new format from V5 onward).
    - Item_ID is the stable key — never change it.
    - Application column: semicolon-separated values e.g. "WB;Cell Culture"
    - Low_Stock column: TRUE / FALSE
    - Current_Stock column: number or blank
    - After running, commit data/inventory.json to GitHub to update the site.
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl is required. Run: pip install openpyxl") from exc

INVENTORY_SHEET = "Inventory"

# All fields in output JSON — order matters for readability
SCHEMA = [
    "Item_ID", "Category", "Subcategory", "Item_Name", "Manufacturer",
    "Cat_No", "CAS_No", "MW_kDa",
    "Application", "Storage", "Location", "Sub_Location",
    "Quantity_Size", "Order_Unit", "Current_Stock", "Low_Stock",
    "Opened_Date", "Expiry_Date", "Requester", "MSDS_URL", "Note",
]

# Valid values for key fields — warn if something unexpected appears
VALID_CATEGORIES = {
    "Chemical", "Antibody/Protein", "Products"
}
VALID_STORAGE = {"RT", "4°C", "-20°C", "-80°C", "LN2", ""}

STORAGE_NORM = {
    "4'c": "4°C", "4c": "4°C", "4℃": "4°C", "2-8°c": "4°C", "2-8c": "4°C",
    "-20'c": "-20°C", "-20c": "-20°C", "-20℃": "-20°C",
    "-80'c": "-80°C", "-80c": "-80°C", "-80℃": "-80°C",
    "rt": "RT",
}


def clean(value) -> str:
    """Convert any cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y.%m.%d")
    s = str(value).strip()
    return "" if s in ("nan", "NaT", "NaN", "None") else s


def norm_storage(s: str) -> str:
    return STORAGE_NORM.get(s.lower(), s)


def parse_bool(s: str) -> bool:
    return s.upper() in ("TRUE", "1", "YES", "Y")


def parse_stock(s: str):
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_application(s: str) -> list:
    if not s:
        return []
    return [a.strip() for a in s.split(";") if a.strip()]


def read_inventory_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Sheet '{INVENTORY_SHEET}' is empty.")

    headers = [clean(cell) for cell in rows[0]]

    # Validate required columns present
    required = {"Item_ID", "Category", "Item_Name"}
    missing = required - set(headers)
    if missing:
        raise SystemExit(f"Missing required columns: {missing}. Found: {headers}")

    items = []
    warnings = []

    for row_num, row in enumerate(rows[1:], start=2):
        record = {
            headers[i]: clean(row[i]) if i < len(row) else ""
            for i in range(len(headers))
        }

        # Skip completely empty rows
        if not record.get("Item_ID") and not record.get("Item_Name"):
            continue

        # Warn on unexpected category values
        cat = record.get("Category", "")
        if cat and cat not in VALID_CATEGORIES:
            warnings.append(f"  Row {row_num} [{record.get('Item_ID','')}]: unknown Category '{cat}'")

        # Normalise storage
        storage = norm_storage(record.get("Storage", ""))
        if storage and storage not in VALID_STORAGE:
            warnings.append(f"  Row {row_num} [{record.get('Item_ID','')}]: unknown Storage '{storage}'")

        item = {}
        for field in SCHEMA:
            raw = record.get(field, "")
            if field == "Application":
                item[field] = parse_application(raw)
            elif field == "Low_Stock":
                item[field] = parse_bool(raw)
            elif field == "Current_Stock":
                item[field] = parse_stock(raw)
            elif field == "Storage":
                item[field] = storage
            else:
                item[field] = raw

        items.append(item)

    return items, warnings


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        raise SystemExit(1)

    input_path  = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    print(f"Reading: {input_path.name}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if INVENTORY_SHEET not in wb.sheetnames:
        raise SystemExit(
            f"Sheet '{INVENTORY_SHEET}' not found.\n"
            f"Available sheets: {wb.sheetnames}"
        )

    items, warnings = read_inventory_sheet(wb[INVENTORY_SHEET])

    if warnings:
        print(f"\n⚠  {len(warnings)} warnings:")
        for w in warnings:
            print(w)

    cat_counts = dict(Counter(i["Category"] for i in items))

    payload = {
        "meta": {
            "title": "L1521 Inventory",
            "source_file": input_path.name,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "version": "web-v2",
            "record_count": len(items),
            "categories": cat_counts,
            "schema": SCHEMA,
        },
        "items": items,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"\n✓  Wrote {len(items)} items to {output_path}")
    print(f"   Categories: {cat_counts}")
    print(f"\n   Next step: commit {output_path} to GitHub\n")


if __name__ == "__main__":
    main()
